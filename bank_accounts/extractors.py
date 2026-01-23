"""
Bank statement extractors that output standardized CSV format.

CSV Schema:
    date,value_date,narration,debit_amount,credit_amount,reference_number,closing_balance
    2024-01-15,2024-01-15,"UPI-SWIGGY",500.00,0.00,REF123456,45000.00
"""
import csv
import io
import re
from datetime import datetime
from decimal import Decimal
from pathlib import Path

# Excel parsing imports
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import msoffcrypto
    MSOFFCRYPTO_AVAILABLE = True
except ImportError:
    MSOFFCRYPTO_AVAILABLE = False

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False


# Standard CSV columns for bank transactions
BANK_CSV_COLUMNS = [
    'date', 'value_date', 'narration', 'debit_amount',
    'credit_amount', 'reference_number', 'closing_balance'
]


def format_date(d):
    """Format date as YYYY-MM-DD."""
    if d is None:
        return ''
    return d.strftime('%Y-%m-%d')


def format_decimal(d):
    """Format decimal as NNNN.NN."""
    if d is None:
        return '0.00'
    return f'{d:.2f}'


def parse_amount(amount_str):
    """Parse amount string to Decimal."""
    if amount_str is None:
        return Decimal('0.00')
    cleaned = str(amount_str).strip().replace(',', '')
    if not cleaned or cleaned == '' or cleaned == '-':
        return Decimal('0.00')
    try:
        return Decimal(cleaned)
    except Exception:
        return Decimal('0.00')


def parse_date_flexible(date_val):
    """Parse date from various formats (string or datetime object)."""
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        return date_val.date()
    if hasattr(date_val, 'date'):  # datetime-like object
        return date_val.date()

    date_str = str(date_val).strip()

    # Try different formats
    formats = ['%d/%m/%y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d-%m-%y']
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue

    raise ValueError(f"Unable to parse date: {date_val}")


def transactions_to_csv(transactions):
    """Convert list of transaction dicts to CSV string."""
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

    # Write header
    writer.writerow(BANK_CSV_COLUMNS)

    # Write data rows
    for txn in transactions:
        writer.writerow([
            format_date(txn.get('date')),
            format_date(txn.get('value_date')),
            txn.get('narration', ''),
            format_decimal(txn.get('debit_amount', Decimal('0.00'))),
            format_decimal(txn.get('credit_amount', Decimal('0.00'))),
            txn.get('reference_number', ''),
            format_decimal(txn.get('closing_balance', Decimal('0.00'))),
        ])

    return output.getvalue()


# ============================================================================
# TXT/CSV Extractor (HDFC format)
# ============================================================================

def extract_hdfc_txt(file_path, password=None):
    """
    Extract transactions from HDFC-style TXT/CSV file.

    Expected format (comma-separated):
    date,narration,value_date,debit,credit,reference,balance
    DD/MM/YY,...

    Returns: CSV string in standardized format
    """
    file_path = Path(file_path)
    transactions = []

    with open(file_path, 'r') as f:
        lines = f.readlines()

    for i, line in enumerate(lines):
        if i == 0:  # Skip header
            continue
        line = line.strip()
        if not line:
            continue

        parts = line.split(',')
        if len(parts) < 7:
            continue

        try:
            date_str = parts[0].strip()
            if not re.match(r'\d{2}/\d{2}/\d{2}', date_str):
                continue

            date = datetime.strptime(date_str, '%d/%m/%y').date()
            narration = parts[1].strip()
            value_date = datetime.strptime(parts[2].strip(), '%d/%m/%y').date()
            debit_amount = parse_amount(parts[3])
            credit_amount = parse_amount(parts[4])
            reference_number = parts[5].strip()
            closing_balance = parse_amount(parts[6])

            # Skip rows with no transaction amounts
            if debit_amount == 0 and credit_amount == 0:
                continue

            transactions.append({
                'date': date,
                'narration': narration,
                'value_date': value_date,
                'debit_amount': debit_amount,
                'credit_amount': credit_amount,
                'reference_number': reference_number,
                'closing_balance': closing_balance,
            })
        except (ValueError, IndexError):
            continue

    return transactions_to_csv(transactions)


# ============================================================================
# XLSX Extractor (ICICI format)
# ============================================================================

def load_xlsx_workbook(file_path, password=None):
    """Load an xlsx workbook, handling encryption if needed."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for xlsx parsing. Install with: uv add openpyxl")

    file_path = Path(file_path)

    # Try direct loading first
    try:
        return load_workbook(file_path, data_only=True)
    except Exception:
        pass

    # Check if encrypted
    if not MSOFFCRYPTO_AVAILABLE:
        raise ImportError("msoffcrypto-tool is required for encrypted xlsx files. Install with: uv add msoffcrypto-tool")

    with open(file_path, 'rb') as f:
        office_file = msoffcrypto.OfficeFile(f)
        if not office_file.is_encrypted():
            raise ValueError(f"Failed to open xlsx file: {file_path}")

        # Try with provided password or empty password
        decrypted = io.BytesIO()
        passwords_to_try = [password, '', None] if password else ['', None]

        for pwd in passwords_to_try:
            try:
                f.seek(0)
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=pwd or '')
                office_file.decrypt(decrypted)
                decrypted.seek(0)
                return load_workbook(decrypted, data_only=True)
            except Exception:
                continue

        raise ValueError("Could not decrypt xlsx file. Set password in the extraction pipeline or use --password option.")


def find_header_row(sheet):
    """Find the header row in an xlsx sheet by looking for common column names."""
    header_patterns = ['date', 'narration', 'debit', 'credit', 'balance', 'description', 'particulars']

    for row_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), 1):
        if row is None:
            continue
        row_lower = [str(cell).lower() if cell else '' for cell in row]
        matches = sum(1 for pattern in header_patterns if any(pattern in cell for cell in row_lower))
        if matches >= 3:  # Found at least 3 matching column headers
            return row_idx, row

    return None, None


def map_xlsx_columns(header_row):
    """Map xlsx column indices to transaction fields based on header names."""
    column_map = {}
    header_lower = [str(cell).lower().strip() if cell else '' for cell in header_row]

    # Common column name mappings
    mappings = {
        'date': ['date', 'txn date', 'transaction date', 'trans date', 'txndate'],
        'narration': ['narration', 'description', 'particulars', 'details', 'remarks', 'transaction details'],
        'value_date': ['value date', 'value dt', 'valuedate', 'val date'],
        'debit': ['debit', 'debit amount', 'withdrawal', 'dr', 'debit amt', 'dr amount'],
        'credit': ['credit', 'credit amount', 'deposit', 'cr', 'credit amt', 'cr amount'],
        'reference': ['ref', 'reference', 'chq/ref', 'chq', 'ref no', 'reference no', 'chq/ref number'],
        'balance': ['balance', 'closing balance', 'closing bal', 'running balance', 'available balance'],
    }

    for field, patterns in mappings.items():
        for idx, header in enumerate(header_lower):
            if any(pattern in header for pattern in patterns):
                column_map[field] = idx
                break

    return column_map


def extract_icici_xlsx(file_path, password=None):
    """
    Extract transactions from ICICI-style XLSX file.

    Returns: CSV string in standardized format
    """
    wb = load_xlsx_workbook(file_path, password)
    sheet = wb.active

    # Find header row
    header_row_idx, header_row = find_header_row(sheet)
    if header_row_idx is None:
        raise ValueError('Could not find header row in xlsx file')

    # Map columns
    column_map = map_xlsx_columns(header_row)

    # Parse transactions
    transactions = []
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        try:
            # Extract values using column mapping
            date_val = row[column_map.get('date', 0)] if 'date' in column_map else None
            if date_val is None:
                continue

            date = parse_date_flexible(date_val)
            if date is None:
                continue

            narration = str(row[column_map.get('narration', 1)] or '').strip() if 'narration' in column_map else ''

            value_date_val = row[column_map.get('value_date', 2)] if 'value_date' in column_map else date_val
            try:
                value_date = parse_date_flexible(value_date_val)
            except (ValueError, TypeError):
                value_date = date

            debit_val = row[column_map.get('debit', 3)] if 'debit' in column_map else 0
            credit_val = row[column_map.get('credit', 4)] if 'credit' in column_map else 0
            ref_val = row[column_map.get('reference', 5)] if 'reference' in column_map else ''
            balance_val = row[column_map.get('balance', 6)] if 'balance' in column_map else 0

            # Parse amounts
            debit = parse_amount(debit_val)
            credit = parse_amount(credit_val)
            balance = parse_amount(balance_val)
            reference = str(ref_val or '').strip()

            # Skip rows with no transaction amounts
            if debit == 0 and credit == 0:
                continue

            transactions.append({
                'date': date,
                'narration': narration,
                'value_date': value_date,
                'debit_amount': debit,
                'credit_amount': credit,
                'reference_number': reference,
                'closing_balance': balance,
            })
        except (ValueError, IndexError, TypeError):
            continue

    return transactions_to_csv(transactions)


# ============================================================================
# PDF Extractor (SBI format)
# ============================================================================

def parse_pdf_date(date_str):
    """Parse date from PDF in DD-MM-YY format."""
    if date_str is None:
        return None
    date_str = str(date_str).strip()
    # Try DD-MM-YY format (SBI format)
    try:
        return datetime.strptime(date_str, '%d-%m-%y').date()
    except ValueError:
        pass
    # Try other formats
    formats = ['%d-%m-%Y', '%d/%m/%y', '%d/%m/%Y']
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def is_valid_pdf_transaction_row(row):
    """Check if a PDF row looks like a valid transaction row."""
    if not row or len(row) < 5:
        return False

    # Filter out None values for checking
    non_empty = [c for c in row if c is not None and str(c).strip()]
    if len(non_empty) < 3:
        return False

    # First cell should be a date (DD-MM-YY pattern)
    first_cell = str(row[0]).strip() if row[0] else ''
    if not re.match(r'\d{2}-\d{2}-\d{2}', first_cell):
        return False

    return True


def extract_sbi_pdf(file_path, password=None):
    """
    Extract transactions from SBI-style PDF bank statement.

    Returns: CSV string in standardized format
    """
    if not PDFPLUMBER_AVAILABLE:
        raise ImportError('pdfplumber is required for PDF parsing. Install with: uv add pdfplumber')

    # Try passwords in order: provided, empty
    passwords_to_try = [password, None] if password else [None]

    pdf = None
    for pwd in passwords_to_try:
        try:
            pdf = pdfplumber.open(file_path, password=pwd)
            break
        except Exception:
            continue

    if pdf is None:
        raise ValueError('Failed to open PDF file. Set password in the extraction pipeline or use --password option.')

    transactions = []
    try:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table:
                    continue

                for row in table:
                    if not is_valid_pdf_transaction_row(row):
                        continue

                    try:
                        # Row format varies but last 3 columns are always: Credit, Debit, Balance
                        date_str = str(row[0]).strip()
                        date = parse_pdf_date(date_str)
                        if date is None:
                            continue

                        # Transaction reference (narration) is typically column 1
                        narration = str(row[1] or '').strip()

                        # Use negative indexing - last 3 columns are always Credit, Debit, Balance
                        if len(row) >= 6:
                            credit = parse_amount(row[-3])
                            debit = parse_amount(row[-2])
                            balance = parse_amount(row[-1])
                            # Ref is typically at index -4 or column before Credit
                            ref = str(row[-4] or '').strip() if len(row) >= 7 else ''
                        else:
                            continue

                        # Skip rows with no transaction amounts
                        if credit == 0 and debit == 0:
                            continue

                        transactions.append({
                            'date': date,
                            'narration': narration,
                            'value_date': date,  # PDF doesn't have separate value date
                            'debit_amount': debit,
                            'credit_amount': credit,
                            'reference_number': ref,
                            'closing_balance': balance,
                        })
                    except (ValueError, IndexError, TypeError):
                        continue
    finally:
        pdf.close()

    return transactions_to_csv(transactions)


# ============================================================================
# Extractor Registry
# ============================================================================

EXTRACTORS = {
    'sbi_pdf': extract_sbi_pdf,
    'icici_xlsx': extract_icici_xlsx,
    'hdfc_txt': extract_hdfc_txt,
    'generic_xlsx': extract_icici_xlsx,
    'generic_txt': extract_hdfc_txt,
}


def get_extractor(extractor_name):
    """Get an extractor function by name."""
    return EXTRACTORS.get(extractor_name)


def detect_extractor(file_path):
    """Auto-detect the appropriate extractor based on file extension."""
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()

    if suffix == '.pdf':
        return 'sbi_pdf'
    elif suffix in ['.xlsx', '.xls']:
        return 'generic_xlsx'
    elif suffix in ['.txt', '.csv']:
        return 'generic_txt'
    else:
        return None
