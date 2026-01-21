"""
ICICI Bank XLSX extractor.

Extracts transactions from ICICI bank Excel statements.
Also used as generic_xlsx extractor.
Ported from bank_accs/extractors.py.
"""
import csv
import io
from decimal import Decimal
from datetime import datetime
from typing import Optional
from pathlib import Path

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import msoffcrypto
    MSOFFCRYPTO_AVAILABLE = True
except ImportError:
    MSOFFCRYPTO_AVAILABLE = False

from . import BaseExtractor, ExtractionResult, ArtifactSpec, register_extractor


# Standard CSV columns for bank transactions
BANK_CSV_COLUMNS = [
    'row_id', 'date', 'value_date', 'narration', 'debit_amount',
    'credit_amount', 'reference_number', 'closing_balance'
]


def format_date(d):
    """Format date as YYYY-MM-DD."""
    if d is None:
        return ''
    return d.strftime('%Y-%m-%d')


def format_decimal(d):
    """Format decimal as NNNN.NN."""
    if d is None:
        return '0.00'
    return f'{d:.2f}'


def parse_amount(amount_str):
    """Parse amount string to Decimal."""
    if amount_str is None:
        return Decimal('0.00')
    cleaned = str(amount_str).strip().replace(',', '')
    if not cleaned or cleaned == '' or cleaned == '-':
        return Decimal('0.00')
    try:
        return Decimal(cleaned)
    except Exception:
        return Decimal('0.00')


def parse_date_flexible(date_val):
    """Parse date from various formats (string or datetime object)."""
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        return date_val.date()
    if hasattr(date_val, 'date'):
        return date_val.date()

    date_str = str(date_val).strip()
    formats = ['%d/%m/%y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d-%m-%y']
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue

    raise ValueError(f"Unable to parse date: {date_val}")


def load_xlsx_workbook(file_bytes: bytes, password: Optional[str] = None):
    """Load an xlsx workbook from bytes, handling encryption if needed."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for xlsx parsing. Install with: uv add openpyxl")

    file_io = io.BytesIO(file_bytes)

    # Try direct loading first
    try:
        file_io.seek(0)
        return load_workbook(file_io, data_only=True)
    except Exception:
        pass

    # Check if encrypted
    if not MSOFFCRYPTO_AVAILABLE:
        raise ImportError("msoffcrypto-tool is required for encrypted xlsx files. Install with: uv add msoffcrypto-tool")

    file_io.seek(0)
    office_file = msoffcrypto.OfficeFile(file_io)
    if not office_file.is_encrypted():
        raise ValueError("Failed to open xlsx file")

    # Try with provided password or empty password
    decrypted = io.BytesIO()
    passwords_to_try = [password, '', None] if password else ['', None]

    for pwd in passwords_to_try:
        try:
            file_io.seek(0)
            office_file = msoffcrypto.OfficeFile(file_io)
            office_file.load_key(password=pwd or '')
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            return load_workbook(decrypted, data_only=True)
        except Exception:
            continue

    raise ValueError("Could not decrypt xlsx file. Check password.")


def find_header_row(sheet):
    """Find the header row in an xlsx sheet by looking for common column names."""
    header_patterns = ['date', 'narration', 'debit', 'credit', 'balance', 'description', 'particulars']

    for row_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), 1):
        if row is None:
            continue
        row_lower = [str(cell).lower() if cell else '' for cell in row]
        matches = sum(1 for pattern in header_patterns if any(pattern in cell for cell in row_lower))
        if matches >= 3:
            return row_idx, row

    return None, None


def map_xlsx_columns(header_row):
    """Map xlsx column indices to transaction fields based on header names."""
    column_map = {}
    header_lower = [str(cell).lower().strip() if cell else '' for cell in header_row]

    mappings = {
        'date': ['date', 'txn date', 'transaction date', 'trans date', 'txndate'],
        'narration': ['narration', 'description', 'particulars', 'details', 'remarks', 'transaction details'],
        'value_date': ['value date', 'value dt', 'valuedate', 'val date'],
        'debit': ['debit', 'debit amount', 'withdrawal', 'dr', 'debit amt', 'dr amount'],
        'credit': ['credit', 'credit amount', 'deposit', 'cr', 'credit amt', 'cr amount'],
        'reference': ['ref', 'reference', 'chq/ref', 'chq', 'ref no', 'reference no', 'chq/ref number'],
        'balance': ['balance', 'closing balance', 'closing bal', 'running balance', 'available balance'],
    }

    for field, patterns in mappings.items():
        for idx, header in enumerate(header_lower):
            if any(pattern in header for pattern in patterns):
                column_map[field] = idx
                break

    return column_map


def transactions_to_csv(transactions):
    """Convert list of transaction dicts to CSV string."""
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

    writer.writerow(BANK_CSV_COLUMNS)

    for row_id, txn in enumerate(transactions, start=1):
        writer.writerow([
            row_id,
            format_date(txn.get('date')),
            format_date(txn.get('value_date')),
            txn.get('narration', ''),
            format_decimal(txn.get('debit_amount', Decimal('0.00'))),
            format_decimal(txn.get('credit_amount', Decimal('0.00'))),
            txn.get('reference_number', ''),
            format_decimal(txn.get('closing_balance', Decimal('0.00'))),
        ])

    return output.getvalue()


@register_extractor
class ICICIXLSXExtractor(BaseExtractor):
    """ICICI Bank Excel Statement Extractor."""
    name = 'icici_xlsx'
    version = '1.0'
    domain = 'bank_account'
    supported_extensions = ['.xlsx', '.xls']

    def extract(self, file_bytes: bytes, password: Optional[str] = None) -> ExtractionResult:
        try:
            wb = load_xlsx_workbook(file_bytes, password)
        except ImportError as e:
            return ExtractionResult(error=str(e))
        except Exception as e:
            return ExtractionResult(error=f'Failed to open xlsx file: {str(e)}')

        sheet = wb.active

        # Find header row
        header_row_idx, header_row = find_header_row(sheet)
        if header_row_idx is None:
            return ExtractionResult(error='Could not find header row in xlsx file')

        # Map columns
        column_map = map_xlsx_columns(header_row)

        # Parse transactions
        transactions = []
        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            try:
                date_val = row[column_map.get('date', 0)] if 'date' in column_map else None
                if date_val is None:
                    continue

                date = parse_date_flexible(date_val)
                if date is None:
                    continue

                narration = str(row[column_map.get('narration', 1)] or '').strip() if 'narration' in column_map else ''

                value_date_val = row[column_map.get('value_date', 2)] if 'value_date' in column_map else date_val
                try:
                    value_date = parse_date_flexible(value_date_val)
                except (ValueError, TypeError):
                    value_date = date

                debit_val = row[column_map.get('debit', 3)] if 'debit' in column_map else 0
                credit_val = row[column_map.get('credit', 4)] if 'credit' in column_map else 0
                ref_val = row[column_map.get('reference', 5)] if 'reference' in column_map else ''
                balance_val = row[column_map.get('balance', 6)] if 'balance' in column_map else 0

                debit = parse_amount(debit_val)
                credit = parse_amount(credit_val)
                balance = parse_amount(balance_val)
                reference = str(ref_val or '').strip()

                if debit == 0 and credit == 0:
                    continue

                transactions.append({
                    'date': date,
                    'narration': narration,
                    'value_date': value_date,
                    'debit_amount': debit,
                    'credit_amount': credit,
                    'reference_number': reference,
                    'closing_balance': balance,
                })
            except (ValueError, IndexError, TypeError):
                continue

        if not transactions:
            return ExtractionResult(error='No transactions found in xlsx file')

        csv_content = transactions_to_csv(transactions)
        artifact = ArtifactSpec(
            artifact_type='transactions',
            content=csv_content,
            content_format='csv',
            row_count=len(transactions),
            data_source_target='bank_account_transactions',
            transformer='bank_transactions',  # Pass-through transformer
        )

        return ExtractionResult(artifacts=[artifact])


# Also register as generic_xlsx
@register_extractor
class GenericXLSXExtractor(ICICIXLSXExtractor):
    """Generic Excel Statement Extractor (same as ICICI)."""
    name = 'generic_xlsx'
    version = '1.0'
