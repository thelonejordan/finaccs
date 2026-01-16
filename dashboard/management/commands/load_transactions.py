import fnmatch
import hashlib
import io
import os
import re
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand
from django.utils import timezone
from dotenv import load_dotenv

from dashboard.models import Transaction, FileLoadLog

# Load environment variables
load_dotenv()

# Excel parsing imports
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import msoffcrypto
    MSOFFCRYPTO_AVAILABLE = True
except ImportError:
    MSOFFCRYPTO_AVAILABLE = False

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False


CATEGORY_PATTERNS = {
    'Food Delivery': ['SWIGGY', 'ZOMATO', 'BLINKIT', 'GROFERS'],
    'Transport': ['UBER', 'OLA', 'RAPIDO'],
    'Shopping': ['AMAZON', 'FLIPKART', 'MYNTRA', 'VENUS TRADER'],
    'Medical': ['WELLNESS FOREVER', 'MEDLIFE', 'PHARMACY', 'MEDICAL'],
    'Utilities': ['ELECTRICITY', 'GAS', 'WATER', 'BROADBAND', 'TRAFFIC'],
    'Bank Charges': ['AMB CHRG', 'CHRG INCL GST'],
    'ATM': ['ATW-', 'NWD-'],
    'Salary/Income': ['SALARY'],
    'Interest': ['INTEREST PAID'],
    'Rent': ['RENT'],
    'Self Transfer': ['UPI-JYOTIRMAYA  MAHANTA', 'UPI-JYOTIRMAYA MAHANTA'],
    'Credit Card Payment': ['PAID VIA CRED'],
    'Cafe & Restaurant': ['CAFE', 'HOTEL', 'RESTAURANT', 'MC DONALDS', 'MCDONALDS'],
    'Groceries': ['WHOLE MART', 'GENERAL ST', 'SUPER MARKET', 'MAHALAXMI GENERAL'],
    'Personal Care': ['SALON', 'UNISEX', 'NATURALS'],
    'Legal Services': ['ONLINE LEGAL', 'LEGAL INDIA'],
    'Entertainment': ['ELEPHANT AND CO'],
    'Sports': ['SPORT', 'CHAMPION'],
}


def categorize_transaction(narration):
    narration_upper = narration.upper()
    for category, patterns in CATEGORY_PATTERNS.items():
        for pattern in patterns:
            if pattern in narration_upper:
                return category
    return 'Uncategorized'


def parse_amount(amount_str):
    cleaned = amount_str.strip().replace(',', '')
    if not cleaned or cleaned == '':
        return Decimal('0.00')
    return Decimal(cleaned)


def parse_date(date_str):
    """Parse date string in DD/MM/YY format."""
    date_str = str(date_str).strip()
    return datetime.strptime(date_str, '%d/%m/%y').date()


def parse_date_flexible(date_val):
    """Parse date from various formats (string or datetime object)."""
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        return date_val.date()
    if hasattr(date_val, 'date'):  # datetime-like object
        return date_val.date()

    date_str = str(date_val).strip()

    # Try different formats
    formats = ['%d/%m/%y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d-%m-%y']
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue

    raise ValueError(f"Unable to parse date: {date_val}")


def load_xlsx_workbook(file_path, password=None):
    """Load an xlsx workbook, handling encryption if needed."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for xlsx parsing. Install with: uv add openpyxl")

    file_path = Path(file_path)

    # Try direct loading first
    try:
        return load_workbook(file_path, data_only=True)
    except Exception:
        pass

    # Check if encrypted
    if not MSOFFCRYPTO_AVAILABLE:
        raise ImportError("msoffcrypto-tool is required for encrypted xlsx files. Install with: uv add msoffcrypto-tool")

    with open(file_path, 'rb') as f:
        office_file = msoffcrypto.OfficeFile(f)
        if not office_file.is_encrypted():
            raise ValueError(f"Failed to open xlsx file: {file_path}")

        # Try with provided password, env password, or empty password
        decrypted = io.BytesIO()
        env_password = os.getenv('XLSX_PASSWORD')
        passwords_to_try = []
        if password:
            passwords_to_try.append(password)
        if env_password:
            passwords_to_try.append(env_password)
        passwords_to_try.extend(['', None])

        for pwd in passwords_to_try:
            try:
                f.seek(0)
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=pwd or '')
                office_file.decrypt(decrypted)
                decrypted.seek(0)
                return load_workbook(decrypted, data_only=True)
            except Exception:
                continue

        raise ValueError(f"Could not decrypt xlsx file. Set XLSX_PASSWORD in .env or use --password option.")


def find_header_row(sheet):
    """Find the header row in an xlsx sheet by looking for common column names."""
    header_patterns = ['date', 'narration', 'debit', 'credit', 'balance', 'description', 'particulars']

    for row_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), 1):
        if row is None:
            continue
        row_lower = [str(cell).lower() if cell else '' for cell in row]
        matches = sum(1 for pattern in header_patterns if any(pattern in cell for cell in row_lower))
        if matches >= 3:  # Found at least 3 matching column headers
            return row_idx, row

    return None, None


def map_xlsx_columns(header_row):
    """Map xlsx column indices to transaction fields based on header names."""
    column_map = {}
    header_lower = [str(cell).lower().strip() if cell else '' for cell in header_row]

    # Common column name mappings
    mappings = {
        'date': ['date', 'txn date', 'transaction date', 'trans date', 'txndate'],
        'narration': ['narration', 'description', 'particulars', 'details', 'remarks', 'transaction details'],
        'value_date': ['value date', 'value dt', 'valuedate', 'val date'],
        'debit': ['debit', 'debit amount', 'withdrawal', 'dr', 'debit amt', 'dr amount'],
        'credit': ['credit', 'credit amount', 'deposit', 'cr', 'credit amt', 'cr amount'],
        'reference': ['ref', 'reference', 'chq/ref', 'chq', 'ref no', 'reference no', 'chq/ref number'],
        'balance': ['balance', 'closing balance', 'closing bal', 'running balance', 'available balance'],
    }

    for field, patterns in mappings.items():
        for idx, header in enumerate(header_lower):
            if any(pattern in header for pattern in patterns):
                column_map[field] = idx
                break

    return column_map


def parse_xlsx_transactions(sheet, column_map, header_row_idx):
    """Parse transactions from xlsx sheet rows."""
    transactions = []

    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        try:
            # Extract values using column mapping
            date_val = row[column_map.get('date', 0)] if 'date' in column_map else None
            if date_val is None:
                continue

            date = parse_date_flexible(date_val)
            if date is None:
                continue

            narration = str(row[column_map.get('narration', 1)] or '').strip() if 'narration' in column_map else ''

            value_date_val = row[column_map.get('value_date', 2)] if 'value_date' in column_map else date_val
            try:
                value_date = parse_date_flexible(value_date_val)
            except (ValueError, TypeError):
                value_date = date

            debit_val = row[column_map.get('debit', 3)] if 'debit' in column_map else 0
            credit_val = row[column_map.get('credit', 4)] if 'credit' in column_map else 0
            ref_val = row[column_map.get('reference', 5)] if 'reference' in column_map else ''
            balance_val = row[column_map.get('balance', 6)] if 'balance' in column_map else 0

            # Parse amounts
            debit = parse_amount(str(debit_val) if debit_val else '0')
            credit = parse_amount(str(credit_val) if credit_val else '0')
            balance = parse_amount(str(balance_val) if balance_val else '0')
            reference = str(ref_val or '').strip()

            # Skip rows with no transaction amounts
            if debit == 0 and credit == 0:
                continue

            category = categorize_transaction(narration)

            transactions.append({
                'date': date,
                'narration': narration,
                'value_date': value_date,
                'debit_amount': debit,
                'credit_amount': credit,
                'reference_number': reference,
                'closing_balance': balance,
                'category': category,
            })
        except (ValueError, IndexError, TypeError) as e:
            continue

    return transactions


def compute_file_hash(file_path):
    """Compute SHA-256 hash of a file."""
    sha256_hash = hashlib.sha256()
    with open(file_path, "rb") as f:
        for byte_block in iter(lambda: f.read(4096), b""):
            sha256_hash.update(byte_block)
    return sha256_hash.hexdigest()


def find_matching_pipeline(filename):
    """Find a pipeline that matches the given filename based on file_pattern."""
    from bank_accs.models import ExtractionPipeline

    for pipeline in ExtractionPipeline.objects.all():
        if pipeline.file_pattern and fnmatch.fnmatch(filename, pipeline.file_pattern):
            return pipeline
    return None


def parse_pdf_amount(amount_str):
    """Parse amount from PDF table cell. Returns 0 for '-' or empty values."""
    if amount_str is None:
        return Decimal('0.00')
    amount_str = str(amount_str).strip()
    if not amount_str or amount_str == '-':
        return Decimal('0.00')
    # Remove commas and clean up
    cleaned = amount_str.replace(',', '')
    try:
        return Decimal(cleaned)
    except Exception:
        return Decimal('0.00')


def parse_pdf_date(date_str):
    """Parse date from PDF in DD-MM-YY format."""
    if date_str is None:
        return None
    date_str = str(date_str).strip()
    # Try DD-MM-YY format (SBI format)
    try:
        return datetime.strptime(date_str, '%d-%m-%y').date()
    except ValueError:
        pass
    # Try other formats
    formats = ['%d-%m-%Y', '%d/%m/%y', '%d/%m/%Y']
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def is_valid_pdf_transaction_row(row):
    """Check if a PDF row looks like a valid transaction row."""
    if not row or len(row) < 5:
        return False

    # Filter out None values for checking
    non_empty = [c for c in row if c is not None and str(c).strip()]
    if len(non_empty) < 3:
        return False

    # First cell should be a date (DD-MM-YY pattern)
    first_cell = str(row[0]).strip() if row[0] else ''
    if not re.match(r'\d{2}-\d{2}-\d{2}', first_cell):
        return False

    return True


def parse_pdf_transactions(pdf, password=None):
    """Extract transactions from a PDF bank statement."""
    transactions = []

    for page in pdf.pages:
        tables = page.extract_tables()
        for table in tables:
            if not table:
                continue

            for row in table:
                if not is_valid_pdf_transaction_row(row):
                    continue

                try:
                    # Row format varies but last 3 columns are always: Credit, Debit, Balance
                    # Examples:
                    #   len=7: Date, Narration, None, Ref, Credit, Debit, Balance
                    #   len=8: Date, Narration, None, None, Ref, Credit, Debit, Balance
                    date_str = str(row[0]).strip()
                    date = parse_pdf_date(date_str)
                    if date is None:
                        continue

                    # Transaction reference (narration) is typically column 1
                    narration = str(row[1] or '').strip()

                    # Use negative indexing - last 3 columns are always Credit, Debit, Balance
                    if len(row) >= 6:
                        credit = parse_pdf_amount(row[-3])
                        debit = parse_pdf_amount(row[-2])
                        balance = parse_pdf_amount(row[-1])
                        # Ref is typically at index -4 or column before Credit
                        ref = str(row[-4] or '').strip() if len(row) >= 7 else ''
                    else:
                        continue

                    # Skip rows with no transaction amounts
                    if credit == 0 and debit == 0:
                        continue

                    category = categorize_transaction(narration)

                    transactions.append({
                        'date': date,
                        'narration': narration,
                        'value_date': date,  # PDF doesn't have separate value date
                        'debit_amount': debit,
                        'credit_amount': credit,
                        'reference_number': ref,
                        'closing_balance': balance,
                        'category': category,
                    })
                except (ValueError, IndexError, TypeError):
                    continue

    return transactions


class Command(BaseCommand):
    help = 'Load transactions from bank statement file (.txt, .xlsx, or .pdf)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            help='Path to the bank statement file',
        )
        parser.add_argument(
            '--all',
            action='store_true',
            help='Load all files from bank_accs/data/',
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Clear existing transactions before loading',
        )
        parser.add_argument(
            '--force',
            action='store_true',
            help='Force reload even if file hash matches',
        )
        parser.add_argument(
            '--password',
            type=str,
            help='Password for encrypted xlsx files (can also be set via XLSX_PASSWORD env var)',
        )

    def handle(self, *args, **options):
        file_path = options.get('file')
        load_all = options.get('all')
        password = options.get('password')
        self.force = options.get('force', False)

        # Sync source files from disk to database
        from bank_accs.views import sync_source_files
        sync_source_files()

        if options['clear']:
            deleted_count, _ = Transaction.objects.all().delete()
            log_count, _ = TransactionLog.objects.all().delete()
            self.stdout.write(self.style.WARNING(f'Deleted {deleted_count} transactions and {log_count} log entries'))

        if load_all:
            self.load_all_files(password)
            return

        if not file_path:
            data_dir = Path('bank_accs/data')
            # Look for txt, xlsx, and pdf files
            files = list(data_dir.glob('*.txt')) + list(data_dir.glob('*.xlsx')) + list(data_dir.glob('*.xls')) + list(data_dir.glob('*.pdf'))
            if not files:
                self.stderr.write(self.style.ERROR('No statement files found in bank_accs/data/'))
                return
            file_path = files[0]

        file_path = Path(file_path)
        if not file_path.exists():
            self.stderr.write(self.style.ERROR(f'File not found: {file_path}'))
            return

        transactions_created = self.load_file(file_path, password)
        if transactions_created >= 0:
            self.stdout.write(self.style.SUCCESS(f'Successfully loaded {transactions_created} transactions'))

    def load_all_files(self, password=None):
        """Load all files from bank_accs/data/"""
        data_dir = Path('bank_accs/data')
        files = list(data_dir.glob('*.txt')) + list(data_dir.glob('*.xlsx')) + list(data_dir.glob('*.xls')) + list(data_dir.glob('*.pdf'))

        if not files:
            self.stderr.write(self.style.ERROR('No statement files found in bank_accs/data/'))
            return

        total_created = 0
        files_loaded = 0
        files_skipped = 0
        for file_path in sorted(files):
            transactions_created = self.load_file(file_path, password)
            if transactions_created >= 0:
                total_created += transactions_created
                files_loaded += 1
            else:
                files_skipped += 1

        self.stdout.write(self.style.SUCCESS(
            f'Loaded {total_created} transactions from {files_loaded} files'
            + (f' (skipped {files_skipped} unchanged)' if files_skipped else '')
        ))

    def load_file(self, file_path, password=None):
        """Load transactions from a single file. Returns -1 if skipped."""
        file_path = Path(file_path)
        filename = file_path.name

        # Find or create the source file record
        from bank_accs.models import SourceFile
        source_file, _ = SourceFile.objects.get_or_create(filename=filename)

        # Find matching pipeline for this file
        pipeline = source_file.pipeline or find_matching_pipeline(filename)
        if pipeline and not source_file.pipeline:
            source_file.pipeline = pipeline
            source_file.save(update_fields=['pipeline'])
            self.stdout.write(f'  Matched pipeline: {pipeline.name}')

        # Use pipeline password if no password provided
        if not password and pipeline and pipeline.password:
            password = pipeline.password
            self.stdout.write(f'  Using password from pipeline: {pipeline.name}')

        # Use pipeline's default bank account if source file not linked
        bank_account = source_file.bank_account
        if not bank_account and pipeline and pipeline.default_bank_account:
            bank_account = pipeline.default_bank_account
            source_file.bank_account = bank_account
            source_file.save(update_fields=['bank_account'])
            self.stdout.write(f'  Auto-linked to account: {bank_account.nickname} (from pipeline)')

        # Compute file hash
        current_hash = compute_file_hash(file_path)

        # Check if bank account link changed for existing transactions
        existing_txns = Transaction.objects.filter(source_file=source_file)
        bank_account_changed = False
        if existing_txns.exists():
            first_txn = existing_txns.first()
            if first_txn.bank_account != bank_account:
                bank_account_changed = True
                self.stdout.write(f'Bank account link changed for {filename}')

        # Check if file has changed or bank account link changed
        if not self.force and source_file.file_hash == current_hash and not bank_account_changed:
            self.stdout.write(f'Skipping {filename} (unchanged)')
            return -1

        self.stdout.write(f'Loading transactions from {file_path}')
        if bank_account:
            self.stdout.write(f'  Linking to account: {bank_account.nickname}')
        else:
            self.stdout.write(self.style.WARNING(f'  No bank account linked to {filename}'))

        # Delete existing transactions from this source file before reloading
        if source_file.file_hash or bank_account_changed:
            deleted_count = Transaction.objects.filter(source_file=source_file).delete()[0]
            if deleted_count:
                self.stdout.write(f'  Deleted {deleted_count} existing transactions from this file')

        # Determine file type and parse accordingly
        suffix = file_path.suffix.lower()
        if suffix in ['.xlsx', '.xls']:
            transactions_created, category_counts = self.load_xlsx(file_path, password, bank_account, source_file)
        elif suffix == '.pdf':
            transactions_created, category_counts = self.load_pdf(file_path, password, bank_account, source_file)
        else:
            transactions_created, category_counts = self.load_txt(file_path, bank_account, source_file)

        # Update source file hash and timestamp
        source_file.file_hash = current_hash
        source_file.last_loaded_at = timezone.now()
        source_file.save()

        # Create single FileLoadLog entry for this file load
        if transactions_created > 0:
            FileLoadLog.objects.create(
                source_file=source_file,
                bank_account=bank_account,
                transaction_count=transactions_created,
                file_hash=current_hash,
                category_summary=category_counts,
                link_source='pre_existing' if bank_account else 'none',
            )

        self.stdout.write(self.style.SUCCESS(f'  Loaded {transactions_created} transactions'))
        return transactions_created

    def load_txt(self, file_path, bank_account=None, source_file=None):
        """Load transactions from a txt/csv file."""
        with open(file_path, 'r') as f:
            lines = f.readlines()

        transactions_created = 0
        category_counts = {}
        for i, line in enumerate(lines):
            if i == 0:
                continue
            line = line.strip()
            if not line:
                continue

            parts = line.split(',')
            if len(parts) < 7:
                continue

            try:
                date_str = parts[0].strip()
                if not re.match(r'\d{2}/\d{2}/\d{2}', date_str):
                    continue

                date = parse_date(date_str)
                narration = parts[1].strip()
                value_date = parse_date(parts[2].strip())
                debit_amount = parse_amount(parts[3])
                credit_amount = parse_amount(parts[4])
                reference_number = parts[5].strip()
                closing_balance = parse_amount(parts[6])

                category = categorize_transaction(narration)

                Transaction.objects.create(
                    date=date,
                    narration=narration,
                    value_date=value_date,
                    debit_amount=debit_amount,
                    credit_amount=credit_amount,
                    reference_number=reference_number,
                    closing_balance=closing_balance,
                    category=category,
                    bank_account=bank_account,
                    source_file=source_file,
                )
                # Track category counts
                category_counts[category] = category_counts.get(category, 0) + 1
                transactions_created += 1
            except (ValueError, IndexError) as e:
                self.stderr.write(f'Error parsing line {i + 1}: {e}')
                continue

        return transactions_created, category_counts

    def load_xlsx(self, file_path, password=None, bank_account=None, source_file=None):
        """Load transactions from an xlsx file."""
        try:
            wb = load_xlsx_workbook(file_path, password)
        except Exception as e:
            self.stderr.write(self.style.ERROR(f'Failed to open xlsx file: {e}'))
            return 0, {}

        sheet = wb.active

        # Find header row
        header_row_idx, header_row = find_header_row(sheet)
        if header_row_idx is None:
            self.stderr.write(self.style.ERROR('Could not find header row in xlsx file'))
            return 0, {}

        self.stdout.write(f'Found header row at row {header_row_idx}: {header_row}')

        # Map columns
        column_map = map_xlsx_columns(header_row)
        self.stdout.write(f'Column mapping: {column_map}')

        # Parse transactions
        transactions = parse_xlsx_transactions(sheet, column_map, header_row_idx)

        # Save to database
        transactions_created = 0
        category_counts = {}
        for txn_data in transactions:
            try:
                Transaction.objects.create(
                    **txn_data,
                    bank_account=bank_account,
                    source_file=source_file,
                )
                # Track category counts
                category = txn_data.get('category', 'Uncategorized')
                category_counts[category] = category_counts.get(category, 0) + 1
                transactions_created += 1
            except Exception as e:
                self.stderr.write(f'Error saving transaction: {e}')
                continue

        return transactions_created, category_counts

    def load_pdf(self, file_path, password=None, bank_account=None, source_file=None):
        """Load transactions from a PDF bank statement."""
        if not PDFPLUMBER_AVAILABLE:
            self.stderr.write(self.style.ERROR('pdfplumber is required for PDF parsing. Install with: uv add pdfplumber'))
            return 0, {}

        # Try passwords in order: provided, env var, empty
        env_password = os.getenv('PDF_PASSWORD') or os.getenv('XLSX_PASSWORD')
        passwords_to_try = []
        if password:
            passwords_to_try.append(password)
        if env_password:
            passwords_to_try.append(env_password)
        passwords_to_try.append(None)  # Try without password

        pdf = None
        for pwd in passwords_to_try:
            try:
                pdf = pdfplumber.open(file_path, password=pwd)
                break
            except Exception:
                continue

        if pdf is None:
            self.stderr.write(self.style.ERROR(
                f'Failed to open PDF file. Set PDF_PASSWORD in .env or use --password option.'
            ))
            return 0, {}

        try:
            self.stdout.write(f'  PDF has {len(pdf.pages)} pages')

            # Parse transactions from PDF
            transactions = parse_pdf_transactions(pdf)

            self.stdout.write(f'  Found {len(transactions)} transactions in PDF')

            # Save to database
            transactions_created = 0
            category_counts = {}
            for txn_data in transactions:
                try:
                    Transaction.objects.create(
                        **txn_data,
                        bank_account=bank_account,
                        source_file=source_file,
                    )
                    category = txn_data.get('category', 'Uncategorized')
                    category_counts[category] = category_counts.get(category, 0) + 1
                    transactions_created += 1
                except Exception as e:
                    self.stderr.write(f'Error saving transaction: {e}')
                    continue

            return transactions_created, category_counts
        finally:
            pdf.close()
